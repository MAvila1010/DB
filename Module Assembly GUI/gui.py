import tkinter as tk
from tkinter import messagebox
import openpyxl

class ModuleAssemblyGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Module Assembly")

        # Configure root window to expand
        self.root.rowconfigure(0, weight=1)
        self.root.rowconfigure(1, weight=1)
        self.root.columnconfigure(0, weight=1)
        self.root.columnconfigure(1, weight=1)
        
        # Frame for red squares
        self.squares_frame = tk.Frame(root)
        self.squares_frame.grid(row=1, column=0, padx=10, pady=10, sticky="nsew")
        self.squares_frame.rowconfigure(tuple(range(4)), weight=1)
        self.squares_frame.columnconfigure(tuple(range(16)), weight=1)
        
        self.squares = []
        for i in range(4):  # 4 rows
            row = []
            for j in range(16):  # 16 columns
                square = tk.Label(self.squares_frame, bg='red2', width=4, height=2, relief="ridge")
                square.grid(row=i, column=j, padx=5, pady=5, sticky="nsew")
                row.append(square)
            self.squares.append(row)
        
        # Frame for input boxes
        self.inputs_frame = tk.Frame(root)
        self.inputs_frame.grid(row=1, column=1, padx=10, pady=10, sticky="nsew")
        self.inputs_frame.rowconfigure(tuple(range(35)), weight=1)  # Add extra rows for titles and filename input
        self.inputs_frame.columnconfigure(0, weight=1)
        self.inputs_frame.columnconfigure(1, weight=1)
        
        # Title label for Tile Board
        self.title_label = tk.Label(self.inputs_frame, text="Tile Board")
        self.title_label.grid(row=0, column=0, columnspan=2, padx=5, pady=2)

        # Input box for filename
        self.filename_var = tk.StringVar()
        self.filename_entry = tk.Entry(self.inputs_frame, textvariable=self.filename_var)
        self.filename_entry.grid(row=1, column=0, columnspan=2, padx=5, pady=2, sticky="ew")

        # Title label for Tiles
        self.tiles_label = tk.Label(self.inputs_frame, text="Tiles")
        self.tiles_label.grid(row=2, column=0, columnspan=2, padx=5, pady=2)

        self.input_vars = []
        self.inputs = []
        for i in range(64):
            var = tk.StringVar()
            var.trace_add('write', lambda name, index, mode, var=var, idx=i: self.update_square(var, idx))
            entry = tk.Entry(self.inputs_frame, textvariable=var)
            entry.grid(row=i % 32 + 3, column=i // 32, padx=5, pady=2, sticky="nsew")
            self.input_vars.append(var)
            self.inputs.append(entry)
        
        # Export button
        self.export_button = tk.Button(root, text="Export to Excel", command=self.export_to_excel)
        self.export_button.grid(row=2, column=1, pady=(10, 0), sticky="ew")

        # Clear Entries button
        self.clear_button = tk.Button(root, text="Clear Entries", command=self.clear_entries)
        self.clear_button.grid(row=2, column=0, pady=(10, 0), sticky="ew")

    def update_square(self, var, idx):
        row = idx // 16
        col = idx % 16
        if var.get():
            self.squares[row][col]['bg'] = 'green3'
        else:
            self.squares[row][col]['bg'] = 'red2'

    def export_to_excel(self):

        #Check if all input boxes are filled
        for var in self.input_vars:
            if not var.get():
                messagebox.showerror("Error", "Missing Tiles")
                return
            
        # Get the filename
        filename = self.filename_var.get().strip()
        if not filename:
            messagebox.showerror("Error", "Please enter a filename.")
            return

        # Create a new workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        # Write the filename in the first column of each row
        for i in range(64):
            ws.cell(row=1, column=1, value=filename)

        # Write input box contents to the second column in Excel
        for i, var in enumerate(self.input_vars):
            ws.cell(row=i + 1, column=2, value=var.get())

        # Save workbook
        filename += ".xlsx"  # Append ".xlsx" extension
        wb.save(filename)
        messagebox.showinfo("Export Successful", f"Data has been exported to {filename}")

    def clear_entries(self):
        for var in self.input_vars:
            var.set("")
        self.filename_var.set("")

if __name__ == "__main__":
    root = tk.Tk()
    app = ModuleAssemblyGUI(root)
    root.mainloop()
